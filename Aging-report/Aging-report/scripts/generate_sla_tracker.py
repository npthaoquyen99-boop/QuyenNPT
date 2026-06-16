#!/usr/bin/env python3
"""
Generate or update SLA_Tracker.xlsx from AR report data.
- Merges new AR data into existing tracker (preserves email/response tracking)
- Calculates Payment SLA (NET30) and PIC Response SLA (3 working days)

Usage: python3 generate_sla_tracker.py <ar_xlsx> <tracker_xlsx_path>
  ar_xlsx         : source AR report file
  tracker_xlsx_path: path to save/update tracker (creates new if not exists)
"""

import sys, json, subprocess
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl",
                    "--break-system-packages", "-q"])
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

SCRIPT_DIR = str(Path(__file__).parent) + "/"

# ── SLA config ────────────────────────────────────────────────────────────────
PAYMENT_SLA_DAYS = 30       # NET30
PIC_RESPONSE_SLA_DAYS = 3   # working days

# ── Helpers ───────────────────────────────────────────────────────────────────
def s(style="thin"):
    return Side(style=style, color="BFBFBF")

def border():
    b = s()
    return Border(left=b, right=b, top=b, bottom=b)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def cell_style(c, bold=False, bg=None, fg="000000", h_align="center",
               v_align="center", wrap=False, num_fmt=None, size=10):
    c.font = Font(name="Arial", bold=bold, color=fg, size=size)
    if bg:
        c.fill = fill(bg)
    c.alignment = Alignment(horizontal=h_align, vertical=v_align, wrap_text=wrap)
    c.border = border()
    if num_fmt:
        c.number_format = num_fmt

def add_working_days(start_date, days):
    """Add N working days (Mon-Fri) to a date."""
    if isinstance(start_date, str):
        try:
            start_date = datetime.strptime(start_date, "%d/%m/%Y")
        except Exception:
            return None
    current = start_date
    added = 0
    while added < days:
        current += timedelta(days=1)
        if current.weekday() < 5:  # Mon-Fri
            added += 1
    return current

def working_days_between(start, end):
    """Count working days between two dates."""
    if not start or not end:
        return None
    if isinstance(start, str):
        start = datetime.strptime(start, "%d/%m/%Y")
    if isinstance(end, str):
        end = datetime.strptime(end, "%d/%m/%Y")
    days = 0
    current = start
    while current < end:
        current += timedelta(days=1)
        if current.weekday() < 5:
            days += 1
    return days

def payment_sla_status(over_day):
    """Return (status_label, color_hex) for payment SLA."""
    if over_day <= 0:
        return "✅ Chưa đến hạn", "D9EAD3"
    elif over_day <= PAYMENT_SLA_DAYS:
        return f"⚠️ Trễ {over_day} ngày", "FFF2CC"
    else:
        return f"🔴 Vi phạm ({over_day} ngày)", "FF4C4C"

def pic_sla_status(email_sent_str, response_str, today=None):
    """Return (status_label, color_hex) for PIC response SLA."""
    today = today or datetime.now()
    if not email_sent_str:
        return "—  Chưa gửi mail", "F3F3F3"
    try:
        sent = datetime.strptime(email_sent_str, "%d/%m/%Y")
    except Exception:
        return "—  Chưa gửi mail", "F3F3F3"

    if response_str:
        try:
            responded = datetime.strptime(response_str, "%d/%m/%Y")
            wd = working_days_between(sent, responded)
            if wd is not None and wd <= PIC_RESPONSE_SLA_DAYS:
                return f"✅ Đã phản hồi ({wd} ngày)", "D9EAD3"
            else:
                return f"⚠️ Đã phản hồi muộn ({wd} ngày)", "FFE599"
        except Exception:
            pass

    # No response yet — calculate how many working days since sent
    wd_elapsed = working_days_between(sent, today)
    deadline = add_working_days(sent, PIC_RESPONSE_SLA_DAYS)
    days_left = working_days_between(today, deadline) if today < deadline else 0

    if wd_elapsed is None:
        return "—  Chưa gửi mail", "F3F3F3"
    elif wd_elapsed <= PIC_RESPONSE_SLA_DAYS:
        return f"⏳ Còn {days_left} ngày làm việc", "FFF2CC"
    else:
        return f"🔴 Vi phạm ({wd_elapsed} ngày KLV)", "FF4C4C"

# ── Load existing tracker data ────────────────────────────────────────────────
def load_existing_tracker(tracker_path):
    """Load existing email_sent and response_date keyed by invoice_no."""
    existing = {}
    p = Path(tracker_path)
    if not p.exists():
        return existing
    try:
        wb = load_workbook(tracker_path, data_only=True)
        if "Tracking" not in wb.sheetnames:
            return existing
        ws = wb["Tracking"]
        # Find header row
        headers = {}
        for row in ws.iter_rows(min_row=3, max_row=3, values_only=True):
            for i, h in enumerate(row):
                if h:
                    headers[str(h).strip()] = i
        # Read rows
        for row in ws.iter_rows(min_row=4, values_only=True):
            inv_no = row[headers.get("Invoice #", 3)] if "Invoice #" in headers else None
            email_sent = row[headers.get("Ngày gửi mail", 9)] if "Ngày gửi mail" in headers else None
            pic_resp = row[headers.get("Ngày PIC phản hồi", 10)] if "Ngày PIC phản hồi" in headers else None
            notes = row[headers.get("Ghi chú", 12)] if "Ghi chú" in headers else None
            if inv_no:
                key = str(inv_no).strip()
                existing[key] = {
                    "email_sent": email_sent.strftime("%d/%m/%Y") if hasattr(email_sent, "strftime") else (str(email_sent) if email_sent else ""),
                    "pic_response": pic_resp.strftime("%d/%m/%Y") if hasattr(pic_resp, "strftime") else (str(pic_resp) if pic_resp else ""),
                    "notes": notes or "",
                }
    except Exception as e:
        print(f"Warning: could not read existing tracker: {e}", file=sys.stderr)
    return existing

# ── Main ──────────────────────────────────────────────────────────────────────
def generate_tracker(ar_xlsx, tracker_path):
    # Parse AR
    r = subprocess.run(
        [sys.executable, f"{SCRIPT_DIR}parse_ar.py", ar_xlsx],
        capture_output=True, text=True
    )
    if r.returncode != 0:
        raise RuntimeError(f"parse_ar.py failed:\n{r.stderr}")
    data = json.loads(r.stdout)
    report_date = data["report_date"]
    by_pic = data["by_pic"]
    today = datetime.now()

    # Load existing tracking data
    existing = load_existing_tracker(tracker_path)

    wb = Workbook()

    # ── Sheet: Tracking ───────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Tracking"

    # Title
    ws.merge_cells("A1:M1")
    t = ws["A1"]
    t.value = f"SLA TRACKER — AR AGED DEBTOR  |  Báo cáo đến: {report_date}  |  Cập nhật: {today.strftime('%d/%m/%Y')}"
    t.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    t.fill = fill("1F4E79")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.append([])

    HDRS = [
        "PIC", "Mã KH", "Tên KH", "Invoice #", "Ngày HĐ", "Mô tả",
        "Số tiền (VND)", "Quá hạn (ngày)", "Payment SLA",
        "Ngày gửi mail", "Ngày PIC phản hồi", "PIC Response SLA", "Ghi chú"
    ]
    ws.append(HDRS)
    hdr_row = 3
    HDR_COLORS = {
        "Payment SLA": "C6EFCE",
        "Ngày gửi mail": "DDEBF7",
        "Ngày PIC phản hồi": "DDEBF7",
        "PIC Response SLA": "DDEBF7",
        "Ghi chú": "FCE4D6",
    }
    for ci, h in enumerate(HDRS, 1):
        c = ws.cell(row=hdr_row, column=ci)
        bg = HDR_COLORS.get(h, "2E75B6")
        fg = "000000" if h in HDR_COLORS else "FFFFFF"
        cell_style(c, bold=True, bg=bg, fg=fg)
        c.value = h
    ws.row_dimensions[hdr_row].height = 30

    # Sub-note row
    ws.append([""] * 9 + ["← Nhập dd/mm/yyyy", "← Nhập dd/mm/yyyy", "← Tự động tính", "← Ghi chú tùy ý"])
    note_row = ws.max_row
    for ci in range(1, 14):
        c = ws.cell(row=note_row, column=ci)
        c.font = Font(name="Arial", italic=True, size=8, color="7F7F7F")
        c.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    all_rows = []
    for pic, records in by_pic.items():
        for rec in records:
            inv_key = str(rec["invoice_no"]) if rec["invoice_no"] else ""
            prev = existing.get(inv_key, {})
            email_sent = prev.get("email_sent", "")
            pic_resp = prev.get("pic_response", "")
            notes = prev.get("notes", "")

            pay_label, pay_color = payment_sla_status(rec["over_day"])
            pic_label, pic_color = pic_sla_status(email_sent, pic_resp, today)

            row_vals = [
                pic,
                rec["code"],
                rec["name"] or "",
                f"#{rec['invoice_no']}" if rec["invoice_no"] else "N/A",
                rec["invoice_date"] or "",
                (rec["description"] or "").strip(),
                rec["base_amount"],
                rec["over_day"],
                pay_label,
                email_sent,
                pic_resp,
                pic_label,
                notes,
            ]
            all_rows.append((row_vals, pay_color, pic_color))

    for row_vals, pay_color, pic_color in all_rows:
        ws.append(row_vals)
        ri = ws.max_row
        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row=ri, column=ci)
            c.border = border()
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(
                horizontal="right" if ci in [7, 8] else
                "center" if ci in [1, 2, 4, 5, 8] else "left",
                vertical="center"
            )
            if ci == 7:
                c.number_format = "#,##0"
            # SLA status cells: coloured
            if ci == 9:
                c.fill = fill(pay_color)
                c.font = Font(name="Arial", bold=True, size=10,
                              color="FFFFFF" if pay_color == "FF4C4C" else "000000")
                c.alignment = Alignment(horizontal="center", vertical="center")
            if ci == 12:
                c.fill = fill(pic_color)
                c.font = Font(name="Arial", bold=True, size=10,
                              color="FFFFFF" if pic_color == "FF4C4C" else "000000")
                c.alignment = Alignment(horizontal="center", vertical="center")
            # Editable cols: light yellow background
            if ci in [10, 11, 13]:
                c.fill = fill("FFFDE7")

    # Column widths
    col_widths = [12, 10, 12, 12, 13, 36, 18, 14, 24, 16, 18, 28, 24]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = f"A{hdr_row}:{get_column_letter(len(HDRS))}{ws.max_row}"
    ws.freeze_panes = "A5"

    # ── Sheet: Hướng dẫn ──────────────────────────────────────────────────────
    ws_guide = wb.create_sheet("Hướng dẫn")
    guide_lines = [
        ("SLA TRACKER — HƯỚNG DẪN SỬ DỤNG", True, "1F4E79", "FFFFFF", 13),
        ("", False, None, "000000", 10),
        ("CÁCH CẬP NHẬT HÀNG THÁNG", True, "2E75B6", "FFFFFF", 11),
        ("1. Khi gửi email cho PIC: điền ngày vào cột 'Ngày gửi mail' (định dạng dd/mm/yyyy)", False, None, "000000", 10),
        ("2. Khi PIC phản hồi: điền ngày vào cột 'Ngày PIC phản hồi'", False, None, "000000", 10),
        ("3. Cột 'Ghi chú': ghi lý do PIC cung cấp hoặc ghi chú xử lý", False, None, "000000", 10),
        ("4. Tháng sau upload file AR mới → skill tự merge dữ liệu tracking cũ vào", False, None, "000000", 10),
        ("", False, None, "000000", 10),
        ("QUY TẮC SLA", True, "2E75B6", "FFFFFF", 11),
        (f"• Payment SLA (NET{PAYMENT_SLA_DAYS}): invoice quá hạn > {PAYMENT_SLA_DAYS} ngày = vi phạm 🔴", False, None, "000000", 10),
        (f"• PIC Response SLA: gửi mail > {PIC_RESPONSE_SLA_DAYS} ngày làm việc chưa phản hồi = vi phạm 🔴", False, None, "000000", 10),
        ("", False, None, "000000", 10),
        ("MÀU SẮC", True, "2E75B6", "FFFFFF", 11),
        ("🟢 Xanh: Đúng hạn / Chưa đến hạn", False, "D9EAD3", "000000", 10),
        ("🟡 Vàng: Cần chú ý / Còn trong SLA nhưng đang trễ", False, "FFF2CC", "000000", 10),
        ("🔴 Đỏ: Vi phạm SLA", False, "FF4C4C", "FFFFFF", 10),
        ("⬜ Xám: Chưa có dữ liệu", False, "F3F3F3", "000000", 10),
    ]
    for i, (text, bold, bg, fg, sz) in enumerate(guide_lines, 1):
        c = ws_guide.cell(row=i, column=1, value=text)
        c.font = Font(name="Arial", bold=bold, color=fg, size=sz)
        if bg:
            c.fill = fill(bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws_guide.row_dimensions[i].height = 20
    ws_guide.column_dimensions["A"].width = 80

    wb.save(tracker_path)
    return data

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python3 generate_sla_tracker.py <ar.xlsx> <tracker.xlsx>", file=sys.stderr)
        sys.exit(1)
    generate_tracker(sys.argv[1], sys.argv[2])
    print(f"Tracker saved: {sys.argv[2]}")
