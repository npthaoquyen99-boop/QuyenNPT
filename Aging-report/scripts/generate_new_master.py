#!/usr/bin/env python3
"""
Generate Master Excel for TAM_UNG, AP_AGING, PREPAY.
Usage: python3 generate_new_master.py <parsed_json_file> <output.xlsx>
"""
import sys, json
from datetime import datetime
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    import subprocess; subprocess.run([sys.executable,"-m","pip","install","openpyxl","--break-system-packages","-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HDR_FILL  = PatternFill("solid", fgColor="2E75B6")
HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
TOTAL_FILL= PatternFill("solid", fgColor="D6E4F0")
TOTAL_FONT= Font(bold=True, size=11)
BORDER    = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def cell_style(ws, row, col, value, font=None, fill=None, number_format=None, align='left'):
    c = ws.cell(row=row, column=col, value=value)
    c.border = BORDER
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    if font: c.font = font
    if fill: c.fill = fill
    if number_format: c.number_format = number_format
    return c

def write_tam_ung(data, ws):
    headers = ["Mã NV","Mã NCC","Tên nhân viên","Bộ phận","Phòng ban",
               "Ngày chứng từ","Số chứng từ","Diễn giải","Số tiền còn TU","Ngày hoàn ứng"]
    for i, h in enumerate(headers, 1):
        cell_style(ws, 1, i, h, font=HDR_FONT, fill=HDR_FILL, align='center')
    ws.row_dimensions[1].height = 30
    r = 2
    for emp, records in data["by_employee"].items():
        for rec in records:
            vals = [rec["employee_code"], rec["supplier_code"], rec["employee_name"],
                    rec["dept"], rec["division"], rec["invoice_date"], rec["invoice_no"],
                    rec["description"], rec["amount"], rec["reimbursement_date"]]
            for i, v in enumerate(vals, 1):
                fmt = '#,##0' if i == 9 else None
                cell_style(ws, r, i, v, number_format=fmt, align='right' if i==9 else 'left')
            r += 1
        # subtotal row
        total = sum(x["amount"] for x in records)
        cell_style(ws, r, 3, f"Tổng: {emp}", font=TOTAL_FONT, fill=TOTAL_FILL)
        for i in [1,2,4,5,6,7,8,10]: cell_style(ws, r, i, None, fill=TOTAL_FILL)
        cell_style(ws, r, 9, total, font=TOTAL_FONT, fill=TOTAL_FILL, number_format='#,##0', align='right')
        r += 1
    col_widths = [12,8,30,8,12,14,24,40,18,14]
    for i, w in enumerate(col_widths, 1): ws.column_dimensions[ws.cell(1,i).column_letter].width = w

def write_ap_aging(data, ws):
    headers = ["Mã NCC","Tên NCC","Ngày HĐ","Ngày GL","Số HĐ","Diễn giải",
               "Due date","Term","Tổng VND","Sắp đến >7","Sắp đến ≤7",
               "QH 1-30","QH 31-60","QH 61-90","QH >90"]
    for i, h in enumerate(headers, 1):
        cell_style(ws, 1, i, h, font=HDR_FONT, fill=HDR_FILL, align='center')
    ws.row_dimensions[1].height = 30
    r = 2
    for vendor, records in data["by_vendor"].items():
        for rec in records:
            vals = [rec["vendor_code"], rec["vendor_name"], rec["invoice_date"],
                    rec["gl_date"], rec["invoice_no"], rec["description"],
                    rec["due_date"], rec["term"], rec["total_vnd"],
                    rec["upcoming_7plus"], rec["upcoming_7"],
                    rec["overdue_1_30"], rec["overdue_31_60"],
                    rec["overdue_61_90"], rec["overdue_90plus"]]
            for i, v in enumerate(vals, 1):
                fmt = '#,##0' if i >= 9 else None
                cell_style(ws, r, i, v, number_format=fmt, align='right' if i>=9 else 'left')
            r += 1
        total = sum(x["total_vnd"] for x in records)
        for i in range(1, 16): cell_style(ws, r, i, None, fill=TOTAL_FILL)
        cell_style(ws, r, 2, f"Tổng: {vendor}", font=TOTAL_FONT, fill=TOTAL_FILL)
        cell_style(ws, r, 9, total, font=TOTAL_FONT, fill=TOTAL_FILL, number_format='#,##0', align='right')
        r += 1
    col_widths = [8,30,12,12,24,35,12,10,16,14,14,12,12,12,12]
    for i, w in enumerate(col_widths, 1): ws.column_dimensions[ws.cell(1,i).column_letter].width = w

def write_prepay(data, ws):
    headers = ["STT","Mã NCC","Tên NCC","Ngày GL","Số HĐ","Diễn giải",
               "Requester","Số tiền còn lại","Đơn vị tiền","Ngày hoàn trả","Loại NCC"]
    for i, h in enumerate(headers, 1):
        cell_style(ws, 1, i, h, font=HDR_FONT, fill=HDR_FILL, align='center')
    ws.row_dimensions[1].height = 30
    r = 2
    for sup, records in data["by_supplier"].items():
        for rec in records:
            vals = [rec["stt"], rec["supplier_code"], rec["supplier_name"],
                    rec["gl_date"], rec["invoice_no"], rec["description"],
                    rec["requester"], rec["amount"], rec["currency"],
                    rec["reimbursement_date"], rec["supplier_type"]]
            for i, v in enumerate(vals, 1):
                fmt = '#,##0' if i == 8 else None
                cell_style(ws, r, i, v, number_format=fmt, align='right' if i==8 else 'left')
            r += 1
        total = sum(x["amount"] for x in records)
        for i in range(1,12): cell_style(ws, r, i, None, fill=TOTAL_FILL)
        cell_style(ws, r, 3, f"Tổng: {sup}", font=TOTAL_FONT, fill=TOTAL_FILL)
        cell_style(ws, r, 8, total, font=TOTAL_FONT, fill=TOTAL_FILL, number_format='#,##0', align='right')
        r += 1
    col_widths = [5,8,30,12,24,35,35,18,8,14,10]
    for i, w in enumerate(col_widths, 1): ws.column_dimensions[ws.cell(1,i).column_letter].width = w

def generate(json_path, output_path):
    with open(json_path, encoding='utf-8') as f:
        data = json.load(f)
    ft = data.get("file_type", data.get("account",""))
    period = data.get("period","")
    account = data.get("account","")

    wb = openpyxl.Workbook()
    ws = wb.active

    if data.get("by_employee"):
        ws.title = "Tạm Ứng Master"
        ws.append([f"BÁO CÁO TẠM ỨNG - TK{account} - KỲ {period}"])
        ws.cell(1,1).font = Font(bold=True, size=13)
        ws.append([])
        # shift data down by inserting rows
        write_tam_ung(data, ws)
    elif data.get("by_vendor"):
        ws.title = "AP Aging Master"
        ws.append([f"BÁO CÁO AP AGING - TK{account} - KỲ {period}"])
        ws.cell(1,1).font = Font(bold=True, size=13)
        ws.append([])
        write_ap_aging(data, ws)
    elif data.get("by_supplier"):
        ws.title = "Trả Trước Master"
        ws.append([f"BÁO CÁO TRẢ TRƯỚC NCC - TK{account} - KỲ {period}"])
        ws.cell(1,1).font = Font(bold=True, size=13)
        ws.append([])
        write_prepay(data, ws)

    wb.save(output_path)
    print(f"Master saved: {output_path}")

if __name__ == "__main__":
    generate(sys.argv[1], sys.argv[2])
