#!/usr/bin/env python3
"""
Parse AP Aging (TK3311 sheet "Aging", TK33873 sheet "AP22").
Output JSON: {"report_date","account","by_vendor":{name:[records]}}
"""
import sys, json, re
from datetime import datetime
try:
    import openpyxl
except ImportError:
    import subprocess; subprocess.run([sys.executable,"-m","pip","install","openpyxl","--break-system-packages","-q"])
    import openpyxl

def parse_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.strftime("%d/%m/%Y")
    s = str(v).strip()
    m = re.search(r'(\d{4}-\d{2}-\d{2})', s)
    if m:
        try:
            d = datetime.strptime(m.group(1), '%Y-%m-%d')
            return d.strftime('%d/%m/%Y')
        except: pass
    return s

def parse_float(v):
    if v is None or str(v).strip() in ['-','','None']: return 0.0
    try: return float(v)
    except: return 0.0

def is_vendor_header(row):
    """Row is vendor summary if col0 matches 'CODE - Name' pattern and col4 (due date) is empty."""
    v0 = str(row[0] or '').strip()
    return bool(re.match(r'^\d+\s*-\s*.+', v0)) and (row[4] is None or str(row[4]).strip() in ['','None'])

def is_detail_row(row):
    """Detail row has a date in col0."""
    v0 = row[0]
    return isinstance(v0, datetime) or (isinstance(v0, str) and re.search(r'\d{4}-\d{2}-\d{2}', str(v0)))

def parse_ap_aging(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets_upper = {s.upper(): s for s in wb.sheetnames}
    account = "3311"

    # Pick sheet
    if 'AP22' in sheets_upper:
        ws = wb[sheets_upper['AP22']]
        account = "33873"
    elif 'AGING' in sheets_upper:
        ws = wb[sheets_upper['AGING']]
    else:
        ws = wb[wb.sheetnames[0]]

    # Extract report date from header rows
    report_date = "N/A"
    for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
        for cell in row:
            if cell and isinstance(cell, datetime):
                report_date = cell.strftime("%d/%m/%Y")
            elif cell:
                m = re.search(r'(\d{2}/\d{2}/\d{4})', str(cell))
                if m: report_date = m.group(1)

    # Find header row (has "Ngày HĐ" in col0)
    data_start = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        if row[0] and 'ngày hđ' in str(row[0]).lower():
            data_start = i + 1
            break
    if not data_start:
        data_start = 7  # fallback

    # Col indices (fixed position based on format):
    # 0=Ngày HĐ, 1=Ngày GL, 2=Số HĐ, 3=Diễn giải, 4=Due date, 5=Term,
    # 6=Total VND, 7=FX, 8=FX amt, 9=Sắp đến >7, 10=Sắp đến ≤7,
    # 11=QH 1-30, 12=QH 31-60, 13=QH 61-90, 14=QH >90

    records = []
    current_vendor_code = ""
    current_vendor_name = ""

    for row in ws.iter_rows(min_row=data_start, values_only=True):
        v0 = str(row[0] or '').strip()
        if not any(c is not None for c in row): continue
        if 'tổng cộng' in v0.lower() or 'total' in v0.lower(): continue

        if is_vendor_header(row):
            m = re.match(r'^(\d+)\s*-\s*(.+)', v0)
            if m:
                current_vendor_code = m.group(1).strip()
                current_vendor_name = m.group(2).strip()
            continue

        if is_detail_row(row):
            total = parse_float(row[6] if len(row) > 6 else None)
            if total == 0: continue
            records.append({
                "vendor_code":    current_vendor_code,
                "vendor_name":    current_vendor_name,
                "invoice_date":   parse_date(row[0]),
                "gl_date":        parse_date(row[1]),
                "invoice_no":     str(row[2] or '').strip(),
                "description":    str(row[3] or '').strip(),
                "due_date":       parse_date(row[4]),
                "term":           str(row[5] or '').strip(),
                "total_vnd":      total,
                "upcoming_7plus": parse_float(row[9] if len(row) > 9 else None),
                "upcoming_7":     parse_float(row[10] if len(row) > 10 else None),
                "overdue_1_30":   parse_float(row[11] if len(row) > 11 else None),
                "overdue_31_60":  parse_float(row[12] if len(row) > 12 else None),
                "overdue_61_90":  parse_float(row[13] if len(row) > 13 else None),
                "overdue_90plus": parse_float(row[14] if len(row) > 14 else None),
            })

    # Group by vendor
    by_vendor = {}
    for r in records:
        key = f"{r['vendor_code']} - {r['vendor_name']}" if r['vendor_code'] else r['vendor_name']
        if key not in by_vendor: by_vendor[key] = []
        by_vendor[key].append(r)

    return {
        "report_date": report_date,
        "account": account,
        "total_records": len(records),
        "by_vendor": by_vendor
    }

if __name__ == "__main__":
    print(json.dumps(parse_ap_aging(sys.argv[1]), ensure_ascii=False, indent=2))
