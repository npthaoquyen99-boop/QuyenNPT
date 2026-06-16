#!/usr/bin/env python3
"""
Generate SLA Tracker Excel for TAM_UNG, AP_AGING, PREPAY.
Usage: python3 generate_new_sla_tracker.py <parsed_json> <output.xlsx>
"""
import sys, json
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

TODAY = datetime.today()

def s(): return Side(style="thin", color="BFBFBF")
def border(): return Border(left=s(), right=s(), top=s(), bottom=s())

def fill(c): return PatternFill("solid", fgColor=c)

def cs(cell, bold=False, bg=None, fg="000000", align="center", wrap=False, fmt=None, size=10):
    cell.font = Font(name="Arial", bold=bold, color=fg, size=size)
    if bg: cell.fill = fill(bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = border()
    if fmt: cell.number_format = fmt

def parse_date(s):
    if not s: return None
    for fmt in ["%d/%m/%Y", "%Y-%m-%d"]:
        try: return datetime.strptime(s, fmt)
        except: pass
    return None

def days_overdue(date_str):
    d = parse_date(date_str)
    if not d: return None
    delta = (TODAY - d).days
    return delta  # positive = overdue, negative = not yet

def sla_status(days):
    if days is None: return "N/A", "CCCCCC"
    if days > 90: return "Vi phạm >90 ngày", "FF0000"
    if days > 60: return "Vi phạm >60 ngày", "FF4444"
    if days > 30: return "Vi phạm >30 ngày", "FF9900"
    if days > 0:  return f"Quá hạn {days} ngày", "FFCC00"
    if days > -7: return "Sắp đến hạn", "FFF2CC"
    return "Trong hạn", "E2EFDA"

HDR_FILL = "2E75B6"
HDR_FONT_COLOR = "FFFFFF"

def write_header(ws, cols):
    for i, (h, w) in enumerate(cols, 1):
        c = ws.cell(1, i, h)
        cs(c, bold=True, bg=HDR_FILL, fg=HDR_FONT_COLOR, align="center")
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[1].height = 28

def generate(json_path, output_path):
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    ft      = data.get("file_type", "")
    account = data.get("account", "")
    period  = data.get("period", "")

    wb = Workbook()
    ws = wb.active

    if ft == "TAM_UNG":
        ws.title = "SLA Tạm Ứng"
        cols = [
            ("Tên nhân viên", 28), ("Mã NV", 12), ("Bộ phận", 10),
            ("Ngày CT", 14), ("Số CT", 22), ("Diễn giải", 36),
            ("Số tiền còn TU", 16), ("Ngày hoàn ứng", 14),
            ("Số ngày quá hạn", 16), ("Trạng thái SLA", 22)
        ]
        write_header(ws, cols)
        r = 2
        for emp, records in data.get("by_employee", {}).items():
            for rec in records:
                days = days_overdue(rec.get("reimbursement_date"))
                status, color = sla_status(days)
                row = [
                    emp, rec.get("employee_code",""), rec.get("dept",""),
                    rec.get("invoice_date",""), rec.get("invoice_no",""),
                    rec.get("description",""), rec.get("amount",0),
                    rec.get("reimbursement_date",""),
                    days if days is not None else "",
                    status
                ]
                for i, v in enumerate(row, 1):
                    c = ws.cell(r, i, v)
                    fmt = "#,##0" if i == 7 else None
                    bg = color if i == 10 else None
                    cs(c, bg=bg, align="right" if i==7 else "center" if i in [1,2,3,4,8,9,10] else "left", fmt=fmt)
                r += 1

    elif ft == "AP_AGING":
        ws.title = "SLA AP Aging"
        cols = [
            ("Mã NCC", 10), ("Tên NCC", 30), ("Số HĐ", 24),
            ("Diễn giải", 34), ("Ngày HĐ", 12), ("Due date", 12), ("Term", 10),
            ("Tổng VND", 16), ("QH 1-30", 12), ("QH 31-60", 12),
            ("QH 61-90", 12), ("QH >90", 12), ("Trạng thái SLA", 22)
        ]
        write_header(ws, cols)
        r = 2
        for vendor, records in data.get("by_vendor", {}).items():
            for rec in records:
                overdue = sum([
                    rec.get("overdue_1_30",0), rec.get("overdue_31_60",0),
                    rec.get("overdue_61_90",0), rec.get("overdue_90plus",0)
                ])
                days = days_overdue(rec.get("due_date"))
                status, color = sla_status(days if overdue > 0 else (days if days else -999))
                row = [
                    rec.get("vendor_code",""), rec.get("vendor_name",""),
                    rec.get("invoice_no",""), rec.get("description",""),
                    rec.get("invoice_date",""), rec.get("due_date",""), rec.get("term",""),
                    rec.get("total_vnd",0), rec.get("overdue_1_30",0),
                    rec.get("overdue_31_60",0), rec.get("overdue_61_90",0),
                    rec.get("overdue_90plus",0), status
                ]
                for i, v in enumerate(row, 1):
                    c = ws.cell(r, i, v)
                    fmt = "#,##0" if i in [8,9,10,11,12] else None
                    bg = color if i == 13 else None
                    cs(c, bg=bg, align="right" if i in [8,9,10,11,12] else "left", fmt=fmt)
                r += 1

    elif ft == "PREPAY":
        ws.title = "SLA Trả Trước"
        cols = [
            ("Mã NCC", 10), ("Tên NCC", 28), ("Ngày GL", 12),
            ("Số HĐ", 22), ("Diễn giải", 34), ("Requester", 30),
            ("Số tiền", 16), ("Ngày hoàn trả", 14),
            ("Số ngày quá hạn", 16), ("Trạng thái SLA", 22)
        ]
        write_header(ws, cols)
        r = 2
        for sup, records in data.get("by_supplier", {}).items():
            for rec in records:
                days = days_overdue(rec.get("reimbursement_date"))
                status, color = sla_status(days)
                row = [
                    rec.get("supplier_code",""), rec.get("supplier_name",""),
                    rec.get("gl_date",""), rec.get("invoice_no",""),
                    rec.get("description",""), rec.get("requester",""),
                    rec.get("amount",0), rec.get("reimbursement_date",""),
                    days if days is not None else "", status
                ]
                for i, v in enumerate(row, 1):
                    c = ws.cell(r, i, v)
                    fmt = "#,##0" if i == 7 else None
                    bg = color if i == 10 else None
                    cs(c, bg=bg, align="right" if i==7 else "left", fmt=fmt)
                r += 1

    # Title row above header
    ws.insert_rows(1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
    tc = ws.cell(1, 1, f"SLA TRACKER — TK{account} — Kỳ {period}")
    tc.font = Font(name="Arial", bold=True, size=13)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    wb.save(output_path)
    print(f"SLA Tracker saved: {output_path}")

if __name__ == "__main__":
    generate(sys.argv[1], sys.argv[2])
