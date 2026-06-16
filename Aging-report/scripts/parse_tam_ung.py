#!/usr/bin/env python3
"""
Parse Báo cáo tạm ứng (TK1411/1412).
Output JSON: {"report_date","account","by_employee":{name:[records]}}
"""
import sys, json, re
from datetime import datetime
try:
    import openpyxl
except ImportError:
    import subprocess; subprocess.run([sys.executable,"-m","pip","install","openpyxl","--break-system-packages","-q"])
    import openpyxl

def parse_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.strftime("%d/%m/%Y")
    return str(v)

def parse_float(v):
    if v is None: return 0.0
    try: return float(v)
    except: return 0.0

def parse_sheet(ws):
    """Parse one sheet (1411 or 1412), return list of records."""
    # Find header row
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        if row[0] and 'mã nhân viên' in str(row[0]).lower():
            header_row = i
            break
    if not header_row:
        return []

    headers = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]
    # Normalize header names
    def find_col(keywords):
        for idx, h in enumerate(headers):
            if h and any(k.lower() in str(h).lower() for k in keywords):
                return idx
        return None

    c_emp   = find_col(['mã nhân viên','employee code'])
    c_ncc   = find_col(['mã ncc','supplier code'])
    c_name  = find_col(['tên nhân viên','employee name'])
    c_dept  = find_col(['bộ phận','division'])
    c_div   = find_col(['phòng ban','department'])
    c_date  = find_col(['ngày chứng từ','invoice date'])
    c_invno = find_col(['số chứng từ','invoice number'])
    c_desc  = find_col(['diễn giải','description'])
    c_amt   = find_col(['số tiền còn tạm ứng','remaining amount'])
    c_reimb = find_col(['ngày hoàn ứng','reimbursement date'])

    records = []
    for row in ws.iter_rows(min_row=header_row+1, values_only=True):
        v0 = str(row[0] or '').strip()
        if not v0 or 'số dư' in v0.lower() or 'ending balance' in v0.lower():
            continue
        amt = parse_float(row[c_amt] if c_amt is not None else None)
        if amt == 0:
            continue
        records.append({
            "employee_code": v0,
            "supplier_code":  str(row[c_ncc] or '').strip() if c_ncc is not None else '',
            "employee_name":  str(row[c_name] or '').strip() if c_name is not None else '',
            "dept":           str(row[c_dept] or '').strip() if c_dept is not None else '',
            "division":       str(row[c_div] or '').strip() if c_div is not None else '',
            "invoice_date":   parse_date(row[c_date] if c_date is not None else None),
            "invoice_no":     str(row[c_invno] or '').strip() if c_invno is not None else '',
            "description":    str(row[c_desc] or '').strip() if c_desc is not None else '',
            "amount":         amt,
            "reimbursement_date": parse_date(row[c_reimb] if c_reimb is not None else None),
        })
    return records

def parse_tam_ung(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    all_records = []
    accounts = []
    for sname in wb.sheetnames:
        if sname.strip() in ['1411','1412']:
            accounts.append(sname.strip())
            all_records.extend(parse_sheet(wb[sname]))

    # Group by employee
    by_employee = {}
    for r in all_records:
        key = r['employee_name'] or r['employee_code']
        if key not in by_employee:
            by_employee[key] = []
        by_employee[key].append(r)

    return {
        "report_date": "N/A",  # not in file, use period from filename
        "account": "/".join(accounts) or "1412",
        "total_records": len(all_records),
        "by_employee": by_employee
    }

if __name__ == "__main__":
    print(json.dumps(parse_tam_ung(sys.argv[1]), ensure_ascii=False, indent=2))
