#!/usr/bin/env python3
"""
Generate Combined Master Excel from multiple parsed JSONs.
Usage: python3 generate_combined_master.py <combined_json> <output.xlsx>
"""
import sys, json
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HDR_COLORS = {
    "TAM_UNG":  "F59E0B",
    "AP_AGING": "7C3AED",
    "PREPAY":   "DC2626",
}

def bdr():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def cs(cell, bold=False, bg=None, fg="000000", align="left", fmt=None):
    cell.font = Font(name="Arial", bold=bold, color=fg, size=10)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = bdr()
    if fmt:
        cell.number_format = fmt

def fmt_vnd(n):
    try:
        return float(n)
    except:
        return 0.0

def write_section_header(ws, source_file, account, period, ncols=13):
    r = ws.max_row + 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    tc = ws.cell(r, 1)
    tc.value = f"Nguon: {source_file}  |  TK{account}  |  Ky {period}"
    tc.font = Font(name="Arial", bold=True, size=11, color="1E40AF")
    tc.fill = PatternFill("solid", fgColor="DBEAFE")
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 22

def write_col_headers(ws, headers, hdr_color):
    # Always use a fresh row — compute ONCE before the loop
    hdr_row = ws.max_row + 1
    for i, (h, w) in enumerate(headers, 1):
        c = ws.cell(hdr_row, i, h)
        cs(c, bold=True, bg=hdr_color, fg="FFFFFF", align="center")
        ws.column_dimensions[c.column_letter].width = w

def write_tam_ung(ws, records, source):
    headers = [
        ("Nguon", 26), ("Ten NV", 28), ("Ma NV", 12), ("Bo phan", 10),
        ("Ngay CT", 14), ("So CT", 24), ("Dien giai", 36),
        ("So tien con TU", 16), ("Ngay hoan ung", 14),
    ]
    write_col_headers(ws, headers, HDR_COLORS["TAM_UNG"])
    for emp, recs in records.get("by_employee", {}).items():
        for rec in recs:
            r = ws.max_row + 1
            row_data = [
                source, emp, rec.get("employee_code",""), rec.get("dept",""),
                rec.get("invoice_date",""), rec.get("invoice_no",""),
                rec.get("description",""), fmt_vnd(rec.get("amount",0)),
                rec.get("reimbursement_date",""),
            ]
            for i, v in enumerate(row_data, 1):
                c = ws.cell(r, i, v)
                cs(c, fmt="#,##0" if i == 8 else None, align="right" if i == 8 else "left")

def write_ap_aging(ws, records, source):
    headers = [
        ("Nguon", 26), ("Ma NCC", 10), ("Ten NCC", 28), ("So HD", 24),
        ("Dien giai", 34), ("Ngay HD", 12), ("Due date", 12), ("Term", 10),
        ("Tong VND", 16), ("QH 1-30", 12), ("QH 31-60", 12),
        ("QH 61-90", 12), ("QH >90", 12),
    ]
    write_col_headers(ws, headers, HDR_COLORS["AP_AGING"])
    for vendor, recs in records.get("by_vendor", {}).items():
        for rec in recs:
            r = ws.max_row + 1
            row_data = [
                source, rec.get("vendor_code",""), rec.get("vendor_name",""),
                rec.get("invoice_no",""), rec.get("description",""),
                rec.get("invoice_date",""), rec.get("due_date",""), rec.get("term",""),
                fmt_vnd(rec.get("total_vnd",0)), fmt_vnd(rec.get("overdue_1_30",0)),
                fmt_vnd(rec.get("overdue_31_60",0)), fmt_vnd(rec.get("overdue_61_90",0)),
                fmt_vnd(rec.get("overdue_90plus",0)),
            ]
            for i, v in enumerate(row_data, 1):
                c = ws.cell(r, i, v)
                cs(c, fmt="#,##0" if i >= 9 else None, align="right" if i >= 9 else "left")

def write_prepay(ws, records, source):
    headers = [
        ("Nguon", 26), ("Ma NCC", 10), ("Ten NCC", 28), ("Ngay GL", 12),
        ("So HD", 24), ("Dien giai", 34), ("Requester", 30),
        ("So tien", 16), ("Ngay hoan tra", 14),
    ]
    write_col_headers(ws, headers, HDR_COLORS["PREPAY"])
    for sup, recs in records.get("by_supplier", {}).items():
        for rec in recs:
            r = ws.max_row + 1
            row_data = [
                source, rec.get("supplier_code",""), rec.get("supplier_name",""),
                rec.get("gl_date",""), rec.get("invoice_no",""),
                rec.get("description",""), rec.get("requester",""),
                fmt_vnd(rec.get("amount",0)), rec.get("reimbursement_date",""),
            ]
            for i, v in enumerate(row_data, 1):
                c = ws.cell(r, i, v)
                cs(c, fmt="#,##0" if i == 8 else None, align="right" if i == 8 else "left")

WRITERS = {"TAM_UNG": write_tam_ung, "AP_AGING": write_ap_aging, "PREPAY": write_prepay}
SHEET_NAMES = {"TAM_UNG": "Tam Ung", "AP_AGING": "AP Aging", "PREPAY": "Tra Truoc"}
MAX_COLS = {"TAM_UNG": 9, "AP_AGING": 13, "PREPAY": 9}

def generate(json_path, output_path):
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    sources = data.get("sources", [])
    wb = Workbook()
    wb.remove(wb.active)
    sheets_written = {}

    for src in sources:
        ft = src.get("file_type", "")
        source_file = src.get("source_file", "")
        writer = WRITERS.get(ft)
        if not writer:
            continue

        sheet_name = SHEET_NAMES.get(ft, ft)
        if sheet_name not in sheets_written:
            ws = wb.create_sheet(sheet_name)
            sheets_written[sheet_name] = ws
        else:
            ws = sheets_written[sheet_name]
            ws.append([])  # blank separator row

        write_section_header(ws, source_file, src.get("account",""), src.get("period",""), MAX_COLS.get(ft, 13))
        writer(ws, src, source_file)

    if not wb.sheetnames:
        wb.create_sheet("No Data")

    wb.save(output_path)
    print(f"Combined master saved: {output_path}")

if __name__ == "__main__":
    generate(sys.argv[1], sys.argv[2])
