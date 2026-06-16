#!/usr/bin/env python3
"""Combined SLA Tracker from all debt types."""
import sys, json
from datetime import datetime, date
from pathlib import Path
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    import subprocess
    subprocess.run([sys.executable,"-m","pip","install","openpyxl","--break-system-packages","-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def parse_date(s):
    if not s: return None
    for fmt in ("%d/%m/%Y","%Y-%m-%d","%m/%d/%Y","%d-%m-%Y"):
        try: return datetime.strptime(str(s).strip(), fmt).date()
        except: pass
    return None

def overdue_days(d_str, today):
    d = parse_date(d_str)
    return (today - d).days if d else None

def sla_status(od):
    if od is None: return "Khong xac dinh", "D3D3D3"
    if od > 90:  return "Qua han >90 ngay",  "FF0000"
    if od > 60:  return "Qua han 61-90 ngay", "FF4444"
    if od > 30:  return "Qua han 31-60 ngay", "FF9900"
    if od > 0:   return "Qua han 1-30 ngay",  "FFCC00"
    if od > -7:  return "Sap den han (<7 ngay)","FFF2CC"
    return "Trong han", "E2EFDA"

def bdr():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def cs(cell, bold=False, bg=None, fg="000000", align="left", fmt=None):
    cell.font = Font(name="Arial", bold=bold, color=fg, size=10)
    if bg: cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = bdr()
    if fmt: cell.number_format = fmt

TYPE_LABEL = {"AR":"Phai thu (AR)","TAM_UNG":"Tam ung","AP_AGING":"Phai tra (AP)","PREPAY":"Tra truoc"}
HDR_BG = {"AR":"2E75B6","TAM_UNG":"F59E0B","AP_AGING":"7C3AED","PREPAY":"DC2626"}

HEADERS = ["Loai","Nguon file","PIC / NV","Doi tac","Invoice","Ngay CT","Due date","Mo ta","So tien (VND)","Qua han (ngay)","Trang thai SLA"]

def write_headers(ws):
    for i, h in enumerate(HEADERS, 1):
        c = ws.cell(1, i, h)
        cs(c, bold=True, bg="1E3A5F", fg="FFFFFF", align="center")
    ws.row_dimensions[1].height = 22

def add_row(ws, loai, source, pic, entity, invoice, inv_date, due_date, desc, amount, od, col_widths):
    status_lbl, status_color = sla_status(od)
    r = ws.max_row + 1
    data = [loai, source, pic, entity, invoice, inv_date, due_date, desc,
            float(amount or 0), od if od is not None else "", status_lbl]
    for i, v in enumerate(data, 1):
        c = ws.cell(r, i, v)
        fmt = "#,##0" if i == 9 else None
        align = "right" if i in (9, 10) else "left"
        bg = status_color if i == 11 else None
        cs(c, bg=bg, fmt=fmt, align=align)
    col_widths[10] = max(col_widths.get(10, 0), len(str(od or "")))

def set_col_widths(ws):
    widths = [14, 26, 20, 28, 22, 12, 12, 34, 16, 14, 26]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

def generate(json_path, output_path):
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)
    today = date.today()
    sources = data.get("sources", [])
    if not sources and "file_type" in data:
        sources = [data]

    wb = Workbook()
    ws = wb.active
    ws.title = "SLA Tracker"
    write_headers(ws)
    ws.freeze_panes = "A2"

    col_widths = {}
    for src in sources:
        ft = src.get("file_type",""); sf = src.get("source_file","")
        loai = TYPE_LABEL.get(ft, ft)

        if ft == "AR":
            for pic, recs in src.get("by_pic", {}).items():
                rpic = "" if pic in ("Chua phan cong","Chưa phân công") else pic
                for r in recs:
                    due = r.get("due_date") or r.get("invoice_date")
                    od = overdue_days(due, today)
                    add_row(ws, loai, sf, rpic,
                            (r.get("name") or str(r.get("code",""))).strip(),
                            r.get("invoice_no",""), r.get("invoice_date",""), due or "",
                            (r.get("description","") or "").strip(), r.get("base_amount",0), od, col_widths)

        elif ft == "TAM_UNG":
            for emp, recs in src.get("by_employee", {}).items():
                for r in recs:
                    due = r.get("reimbursement_date")
                    od = overdue_days(due, today)
                    add_row(ws, loai, sf, emp, emp,
                            r.get("invoice_no",""), r.get("invoice_date",""), due or "",
                            (r.get("description","") or "").strip(), r.get("amount",0), od, col_widths)

        elif ft == "AP_AGING":
            for vendor, recs in src.get("by_vendor", {}).items():
                for r in recs:
                    due = r.get("due_date"); od = overdue_days(due, today)
                    if od is None:
                        if r.get("overdue_90plus",0): od=91
                        elif r.get("overdue_61_90",0): od=75
                        elif r.get("overdue_31_60",0): od=45
                        elif r.get("overdue_1_30",0): od=15
                    add_row(ws, loai, sf, "", vendor,
                            r.get("invoice_no",""), r.get("invoice_date",""), due or "",
                            (r.get("description","") or "").strip(), r.get("total_vnd",0), od, col_widths)

        elif ft == "PREPAY":
            for sup, recs in src.get("by_supplier", {}).items():
                for r in recs:
                    due = r.get("reimbursement_date"); od = overdue_days(due, today)
                    add_row(ws, loai, sf, r.get("requester",""), sup,
                            r.get("invoice_no",""), r.get("gl_date",""), due or "",
                            (r.get("description","") or "").strip(), r.get("amount",0), od, col_widths)

    set_col_widths(ws)
    ws.auto_filter.ref = f"A1:K{ws.max_row}"
    wb.save(output_path)
    print(f"Combined SLA tracker saved: {output_path}")

if __name__ == "__main__":
    generate(sys.argv[1], sys.argv[2])
