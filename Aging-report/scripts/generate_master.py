#!/usr/bin/env python3
"""
Generate master Excel summary from parsed AR data.
Usage: python3 generate_master.py <path_to_xlsx> <output_path>
"""

import sys
import json
import subprocess

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

SCRIPT_DIR = __file__.replace("generate_master.py", "")

def side():
    return Side(style="thin", color="BFBFBF")

def thin_border():
    s = side()
    return Border(left=s, right=s, top=s, bottom=s)

def hex_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

AGING_COLOR = {
    "Over 180 ngày": "FF4C4C",
    "91-180 ngày":   "FF9933",
    "61-90 ngày":    "FFD966",
    "31-60 ngày":    "FFE599",
    "1-30 ngày":     "FFF2CC",
    "Current":       "D9EAD3",
    "N/A":           "F3F3F3",
}
AGING_FONT = {
    "Over 180 ngày": "FFFFFF",
    "91-180 ngày":   "FFFFFF",
    "61-90 ngày":    "000000",
    "31-60 ngày":    "000000",
    "1-30 ngày":     "000000",
    "Current":       "000000",
    "N/A":           "999999",
}
NUM_FMT = "#,##0"


def style_header_cell(c, bg="2E75B6"):
    c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    c.fill = hex_fill(bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border()


def style_data_cell(c, h_align="left", is_amount=False):
    c.border = thin_border()
    c.font = Font(name="Arial", size=10)
    c.alignment = Alignment(horizontal=h_align, vertical="center")
    if is_amount:
        c.number_format = NUM_FMT


def generate_master(source_xlsx, output_path):
    # Parse source file
    result = subprocess.run(
        [sys.executable, f"{SCRIPT_DIR}parse_ar.py", source_xlsx],
        capture_output=True, text=True
    )
    if result.returncode != 0:
        raise RuntimeError(f"parse_ar.py failed: {result.stderr}")
    data = json.loads(result.stdout)

    report_date = data["report_date"]
    by_pic = data["by_pic"]
    all_recs = [r for recs in by_pic.values() for r in recs]

    wb = Workbook()

    # ── Sheet 1: Tổng hợp ──────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Tổng hợp"

    ws_sum.merge_cells("A1:J1")
    t = ws_sum["A1"]
    t.value = f"AR AGED DEBTOR — TỔNG HỢP THEO PIC  |  Đến ngày: {report_date}"
    t.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    t.fill = hex_fill("1F4E79")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 28
    ws_sum.append([])

    SUM_HDRS = ["PIC", "Số KH", "Số Invoice", "Tổng tiền (VND)",
                "Current", "1-30 ngày", "31-60 ngày", "61-90 ngày", "91-180 ngày", "Over 180 ngày"]
    ws_sum.append(SUM_HDRS)
    for col_i in range(1, len(SUM_HDRS) + 1):
        style_header_cell(ws_sum.cell(row=3, column=col_i))
    ws_sum.row_dimensions[3].height = 32

    for pic, records in by_pic.items():
        row_vals = [
            pic,
            len(set(r["code"] for r in records)),
            len(records),
            sum(r["base_amount"] for r in records),
            sum(r["base_amount"] for r in records if r["aging_bucket"] == "Current"),
            sum(r["base_amount"] for r in records if r["aging_bucket"] == "1-30 ngày"),
            sum(r["base_amount"] for r in records if r["aging_bucket"] == "31-60 ngày"),
            sum(r["base_amount"] for r in records if r["aging_bucket"] == "61-90 ngày"),
            sum(r["base_amount"] for r in records if r["aging_bucket"] == "91-180 ngày"),
            sum(r["base_amount"] for r in records if r["aging_bucket"] == "Over 180 ngày"),
        ]
        ws_sum.append(row_vals)
        r_idx = ws_sum.max_row
        for col_i, _ in enumerate(row_vals, 1):
            style_data_cell(ws_sum.cell(row=r_idx, column=col_i),
                            h_align="center" if col_i <= 3 else "right",
                            is_amount=(col_i >= 4))

    # Total row
    ws_sum.append([])
    totals = ["TỔNG CỘNG",
              len(set(r["code"] for r in all_recs)),
              len(all_recs),
              sum(r["base_amount"] for r in all_recs),
              sum(r["base_amount"] for r in all_recs if r["aging_bucket"] == "Current"),
              sum(r["base_amount"] for r in all_recs if r["aging_bucket"] == "1-30 ngày"),
              sum(r["base_amount"] for r in all_recs if r["aging_bucket"] == "31-60 ngày"),
              sum(r["base_amount"] for r in all_recs if r["aging_bucket"] == "61-90 ngày"),
              sum(r["base_amount"] for r in all_recs if r["aging_bucket"] == "91-180 ngày"),
              sum(r["base_amount"] for r in all_recs if r["aging_bucket"] == "Over 180 ngày"),
              ]
    ws_sum.append(totals)
    tot_idx = ws_sum.max_row
    for col_i, _ in enumerate(totals, 1):
        c = ws_sum.cell(row=tot_idx, column=col_i)
        c.font = Font(name="Arial", bold=True, size=10)
        c.fill = hex_fill("D6E4F0")
        c.border = thin_border()
        c.alignment = Alignment(horizontal="center" if col_i <= 3 else "right", vertical="center")
        if col_i >= 4:
            c.number_format = NUM_FMT

    col_widths_sum = [14, 10, 12, 20, 16, 14, 14, 14, 14, 14]
    for i, w in enumerate(col_widths_sum, 1):
        ws_sum.column_dimensions[get_column_letter(i)].width = w
    ws_sum.freeze_panes = "A4"

    # ── Sheet 2: Chi tiết ──────────────────────────────────────────────────
    ws_det = wb.create_sheet("Chi tiết")

    ws_det.merge_cells("A1:J1")
    t2 = ws_det["A1"]
    t2.value = f"AR AGED DEBTOR — CHI TIẾT  |  Đến ngày: {report_date}"
    t2.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    t2.fill = hex_fill("1F4E79")
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws_det.row_dimensions[1].height = 28
    ws_det.append([])

    DET_HDRS = ["PIC", "Mã KH", "Tên KH", "Invoice #", "Ngày HĐ",
                "Mô tả", "Loại", "Số tiền (VND)", "Quá hạn (ngày)", "Tình trạng"]
    ws_det.append(DET_HDRS)
    for col_i in range(1, len(DET_HDRS) + 1):
        style_header_cell(ws_det.cell(row=3, column=col_i))
    ws_det.row_dimensions[3].height = 32

    for pic, records in by_pic.items():
        for rec in records:
            bucket = rec["aging_bucket"]
            row_vals = [
                pic, rec["code"], rec["name"] or "",
                f"#{rec['invoice_no']}" if rec["invoice_no"] else "N/A",
                rec["invoice_date"] or "",
                (rec["description"] or "").strip(),
                rec["type"] or "",
                rec["base_amount"],
                rec["over_day"],
                bucket,
            ]
            ws_det.append(row_vals)
            r_idx = ws_det.max_row
            for col_i, _ in enumerate(row_vals, 1):
                c = ws_det.cell(row=r_idx, column=col_i)
                c.border = thin_border()
                c.font = Font(name="Arial", size=10)
                c.alignment = Alignment(
                    vertical="center",
                    horizontal="right" if col_i in [8, 9] else
                    "center" if col_i in [1, 2, 4, 5, 7, 9] else "left")
                if col_i == 8:
                    c.number_format = NUM_FMT
                if col_i == 10:
                    c.fill = hex_fill(AGING_COLOR.get(bucket, "FFFFFF"))
                    c.font = Font(name="Arial", bold=True, size=10,
                                  color=AGING_FONT.get(bucket, "000000"))
                    c.alignment = Alignment(horizontal="center", vertical="center")

    ws_det.auto_filter.ref = f"A3:{get_column_letter(len(DET_HDRS))}{ws_det.max_row}"
    det_widths = [12, 10, 12, 12, 13, 36, 16, 20, 16, 18]
    for i, w in enumerate(det_widths, 1):
        ws_det.column_dimensions[get_column_letter(i)].width = w
    ws_det.freeze_panes = "A4"

    # ── Sheet per PIC ──────────────────────────────────────────────────────
    for pic, records in by_pic.items():
        ws_p = wb.create_sheet(f"PIC_{pic}")

        ws_p.merge_cells("A1:H1")
        tp = ws_p["A1"]
        tp.value = f"PIC: {pic}  |  Đến ngày: {report_date}"
        tp.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        tp.fill = hex_fill("375623")
        tp.alignment = Alignment(horizontal="center", vertical="center")
        ws_p.row_dimensions[1].height = 26
        ws_p.append([])

        P_HDRS = ["Mã KH", "Tên KH", "Invoice #", "Ngày HĐ",
                  "Mô tả", "Số tiền (VND)", "Quá hạn (ngày)", "Tình trạng"]
        ws_p.append(P_HDRS)
        for col_i in range(1, len(P_HDRS) + 1):
            style_header_cell(ws_p.cell(row=3, column=col_i), bg="4EA72A")
        ws_p.row_dimensions[3].height = 28

        for rec in records:
            bucket = rec["aging_bucket"]
            row_vals = [
                rec["code"], rec["name"] or "",
                f"#{rec['invoice_no']}" if rec["invoice_no"] else "N/A",
                rec["invoice_date"] or "",
                (rec["description"] or "").strip(),
                rec["base_amount"],
                rec["over_day"],
                bucket,
            ]
            ws_p.append(row_vals)
            r_idx = ws_p.max_row
            for col_i, _ in enumerate(row_vals, 1):
                c = ws_p.cell(row=r_idx, column=col_i)
                c.border = thin_border()
                c.font = Font(name="Arial", size=10)
                c.alignment = Alignment(
                    vertical="center",
                    horizontal="right" if col_i in [6, 7] else
                    "center" if col_i in [1, 3, 4] else "left")
                if col_i == 6:
                    c.number_format = NUM_FMT
                if col_i == 8:
                    c.fill = hex_fill(AGING_COLOR.get(bucket, "FFFFFF"))
                    c.font = Font(name="Arial", bold=True, size=10,
                                  color=AGING_FONT.get(bucket, "000000"))
                    c.alignment = Alignment(horizontal="center", vertical="center")

        # Total row
        last_data = ws_p.max_row
        tot_r = last_data + 1
        ws_p.cell(row=tot_r, column=5).value = "TỔNG"
        ws_p.cell(row=tot_r, column=5).font = Font(name="Arial", bold=True, size=10)
        ws_p.cell(row=tot_r, column=5).alignment = Alignment(horizontal="right")
        fc = ws_p.cell(row=tot_r, column=6)
        fc.value = f"=SUM(F4:F{last_data})"
        fc.number_format = NUM_FMT
        fc.font = Font(name="Arial", bold=True, size=10)
        fc.fill = hex_fill("EBF1DE")
        fc.alignment = Alignment(horizontal="right", vertical="center")

        p_widths = [10, 12, 12, 13, 38, 20, 16, 18]
        for i, w in enumerate(p_widths, 1):
            ws_p.column_dimensions[get_column_letter(i)].width = w
        ws_p.freeze_panes = "A4"

    wb.save(output_path)
    return data  # return parsed data for email generation


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python3 generate_master.py <source.xlsx> <output.xlsx>", file=sys.stderr)
        sys.exit(1)
    generate_master(sys.argv[1], sys.argv[2])
    print(f"Master file saved: {sys.argv[2]}")
