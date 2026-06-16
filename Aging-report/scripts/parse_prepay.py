#!/usr/bin/env python3
"""
Parse Báo cáo công nợ trả trước nhà cung cấp (TK3313).
Output JSON: {"report_date","account","by_supplier":{name:[records]}}
"""
import sys, json, re
from datetime import datetime
try:
    import openpyxl
except ImportError:
    import subprocess; subprocess.run([sys.executable,"-m","pip","install","openpyxl","--break-system-packages","-q"])
    import openpyxl

def parse_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.strftime("%d/%m/%Y")
    s = str(v).strip()
    m = re.search(r'(\d{4}-\d{2}-\d{2})', s)
    if m:
        try: return datetime.strptime(m.group(1),'%Y-%m-%d').strftime('%d/%m/%Y')
        except: pass
    return s

def parse_float(v):
    if v is None: return 0.0
    try: return float(v)
    except: return 0.0

def parse_prepay(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets_upper = {s.upper(): s for s in wb.sheetnames}
    ws = wb[sheets_upper.get('AGING', wb.sheetnames[0])]

    # Find header row
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        if row[0] and 'stt' in str(row[0]).lower():
            header_row = i
            break
    if not header_row: header_row = 4

    headers = list(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))[0]

    def find_col(keywords):
        for idx, h in enumerate(headers):
            if h and any(k.lower() in str(h).lower() for k in keywords):
                return idx
        return None

    c_stt    = find_col(['stt'])
    c_code   = find_col(['supplier code'])
    c_name   = find_col(['supplier name'])
    c_gl     = find_col(['gl date'])
    c_inv    = find_col(['invoice number'])
    c_desc   = find_col(['description'])
    c_req    = find_col(['requester'])
    c_amt    = find_col(['remaining amount'])
    c_curr   = find_col(['currency'])
    c_reimb  = find_col(['reimbursement date'])
    c_stype  = find_col(['supplier type'])

    records = []
    for row in ws.iter_rows(min_row=header_row+1, values_only=True):
        v3 = str(row[3] or '').strip() if len(row) > 3 else ''
        if 'số dư cuối kỳ' in v3.lower() or 'ending balance' in v3.lower():
            continue
        v0 = str(row[0] or '').strip()
        if not v0 or not v0.replace(' ','').isdigit():
            continue
        amt = parse_float(row[c_amt] if c_amt is not None else None)
        if amt == 0: continue

        records.append({
            "stt":                v0,
            "supplier_code":      str(row[c_code] or '').strip() if c_code is not None else '',
            "supplier_name":      str(row[c_name] or '').strip() if c_name is not None else '',
            "gl_date":            parse_date(row[c_gl] if c_gl is not None else None),
            "invoice_no":         str(row[c_inv] or '').strip() if c_inv is not None else '',
            "description":        str(row[c_desc] or '').strip() if c_desc is not None else '',
            "requester":          str(row[c_req] or '').strip() if c_req is not None else '',
            "amount":             amt,
            "currency":           str(row[c_curr] or 'VND').strip() if c_curr is not None else 'VND',
            "reimbursement_date": parse_date(row[c_reimb] if c_reimb is not None else None),
            "supplier_type":      str(row[c_stype] or '').strip() if c_stype is not None else '',
        })

    by_supplier = {}
    for r in records:
        key = r['supplier_name'] or r['supplier_code']
        if key not in by_supplier: by_supplier[key] = []
        by_supplier[key].append(r)

    return {
        "report_date": "N/A",
        "account": "3313",
        "total_records": len(records),
        "by_supplier": by_supplier
    }

if __name__ == "__main__":
    print(json.dumps(parse_prepay(sys.argv[1]), ensure_ascii=False, indent=2))
