#!/usr/bin/env python3
"""
Detect file type from Excel file.
Output JSON: {"type": "AR|TAM_UNG|AP_AGING|PREPAY", "account": "...", "period": "MM/YYYY"}
"""
import sys, json, re, os
try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    import openpyxl


def detect_type(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    sheets = [s.strip() for s in wb.sheetnames]
    fname = os.path.basename(filepath).lower()

    # Extract period from filename
    period = "N/A"
    m = re.search(r"(202[0-9])[._-](\d{2})", fname)
    if m:
        period = f"{m.group(2)}/{m.group(1)}"
    if period == "N/A":
        m = re.search(r"(202[0-9])\.(\d{2})", fname)
        if m:
            period = f"{m.group(2)}/{m.group(1)}"
    if period == "N/A":
        m = re.search(r"[_ ]t(\d{2})\.(\d{2})", fname)
        if m:
            period = f"{m.group(1)}/20{m.group(2)}"

    sheet_upper = [s.upper() for s in sheets]

    if any(s in ["1411", "1412"] for s in sheets):
        accs = [s for s in sheets if s in ["1411", "1412"]]
        return {"type": "TAM_UNG", "account": "-".join(accs), "period": period}

    if "AP22" in sheet_upper or "AP08" in sheet_upper:
        return {"type": "AP_AGING", "account": "33873", "period": period}

    if "3311" in sheets and "Aging" in sheets:
        return {"type": "AP_AGING", "account": "3311", "period": period}

    for sname in sheets[:2]:
        ws = wb[sname]
        for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
            for cell in row:
                cs = str(cell or "").lower()
                if "supplier code" in cs:
                    return {"type": "PREPAY", "account": "3313", "period": period}
                if "mã nhân viên" in cs or "employee code" in cs:
                    accs = [s for s in sheets if s in ["1411", "1412"]] or ["1412"]
                    return {"type": "TAM_UNG", "account": "-".join(accs), "period": period}
                if "customer name" in cs and "invoice" in cs:
                    return {"type": "AR", "account": "AR", "period": period}

    if "tam ung" in fname or "tamung" in fname:
        return {"type": "TAM_UNG", "account": "1412", "period": period}
    if "3313" in fname:
        return {"type": "PREPAY", "account": "3313", "period": period}
    if "3311" in fname:
        return {"type": "AP_AGING", "account": "3311", "period": period}
    if "33873" in fname:
        return {"type": "AP_AGING", "account": "33873", "period": period}
    # AR: tên file chứa "ar", "1311", hoặc "bao cao tk13"
    if re.search(r'\bar\b|_ar\d|ar_|ar\d', fname) or "1311" in fname or "bao cao tk13" in fname:
        return {"type": "AR", "account": "AR", "period": period}

    return {"type": "UNKNOWN", "account": "?", "period": period}


if __name__ == "__main__":
    print(json.dumps(detect_type(sys.argv[1]), ensure_ascii=False))
