#!/usr/bin/env python3
"""
Parse AR Aged Debtor Report Excel file and output JSON grouped by PIC.
Usage: python3 parse_ar.py <path_to_xlsx>
"""

import sys
import json
import re
from datetime import datetime

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])
    import openpyxl


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    return str(val)


def get_aging_bucket(row_dict):
    buckets = [
        ("Over 180 ngay",  row_dict.get("over_180", 0) or 0),
        ("91-180 ngay",    row_dict.get("91_180", 0) or 0),
        ("61-90 ngay",     row_dict.get("61_90", 0) or 0),
        ("31-60 ngay",     row_dict.get("31_60", 0) or 0),
        ("1-30 ngay",      row_dict.get("1_30", 0) or 0),
        ("Current",        row_dict.get("current", 0) or 0),
    ]
    for label, amt in buckets:
        if amt and amt > 0:
            return label, amt
    return "N/A", 0


def parse_sheet(ws):
    """Parse mot sheet AR, tra ve (records, report_date)."""
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row[0] == 'Code':
            header_row_idx = i
            break
    if header_row_idx is None:
        return [], None

    report_date = None
    for row in ws.iter_rows(min_row=1, max_row=header_row_idx - 1, values_only=True):
        cell = str(row[0] or "")
        match = re.search(r'den ngay[:\s]+(\d{2}/\d{2}/\d{4})', cell, re.IGNORECASE)
        if not match:
            match = re.search(r'\d{2}/\d{2}/\d{4}', cell)
        if match:
            report_date = match.group(0) if not match.lastindex else match.group(1)
            break

    headers = list(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))[0]
    col_map = {}
    for idx, h in enumerate(headers):
        if h:
            col_map[str(h).strip()] = idx

    def col(*candidates):
        for name in candidates:
            if name in col_map:
                return col_map[name]
        return None

    def get(row, *candidates):
        idx = col(*candidates)
        return row[idx] if idx is not None else None

    pic_col = col("PIC", "Pic", "pic", "Nguoi phu trach", "PIC Name", "Sale Staff", "SaleStaff")

    records = []
    for row in ws.iter_rows(min_row=header_row_idx + 2, values_only=True):
        code = get(row, 'Code', 'CODE', 'Ma KH')
        if code is None or not isinstance(code, (int, float)) or code == 0:
            continue
        base_amount = get(row, 'Total of Base Amount', 'Base Amount', 'Amount', 'So tien') or 0

        record = {
            "code":            int(code),
            "name":            get(row, 'Customer Name', 'Ten KH', 'Name'),
            "transaction":     get(row, 'Transaction Number', 'Trans No', 'Transaction No'),
            "description":     get(row, 'Description', 'Dien giai', 'Desc'),
            "invoice_no":      get(row, 'Transaction Number', 'Trans No', 'Transaction No', 'Invoice', 'Invoice No', 'So HD'),
            "kind":            get(row, 'Kind of Customer', 'Kind', 'Loai KH'),
            "invoice_date":    parse_date(get(row, 'Invoice Date', 'Ngay HD', 'Inv Date')),
            "due_date":        parse_date(get(row, 'Date of payment', 'Due Date', 'Ngay den han', 'Payment Date')),
            "over_day":        int(get(row, 'Over day', 'Overdue Days', 'Over Day') or 0),
            "original_amount": float(get(row, 'Total of Original Amount', 'Original Amount') or 0),
            "base_amount":     float(base_amount),
            "current":         float(get(row, 'Currentliability', 'Current', 'Current Liability') or 0),
            "1_30":            float(get(row, '1 - 30 days', '1-30 days', '1-30') or 0),
            "31_60":           float(get(row, '31 - 60 days', '31-60 days', '31-60') or 0),
            "61_90":           float(get(row, '61 - 90 days', '61-90 days', '61-90') or 0),
            "91_180":          float(get(row, '91 - 180 days', '91-180 days', '91-180') or 0),
            "over_180":        float(get(row, 'Over 180 days', 'Over 180', '>180 days') or 0),
            "type":            get(row, 'Type', 'Loai'),
            "pic":             row[pic_col] if pic_col is not None else None,
        }
        bucket_label, bucket_amt = get_aging_bucket(record)
        record["aging_bucket"] = bucket_label
        record["aging_amount"] = bucket_amt
        records.append(record)

    return records, report_date


def parse_ar_file(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Doc sheet co "AR10" hoac "Aging" trong ten; fallback sheet active
    target_sheets = [s for s in wb.sheetnames if "AR10" in s or "Aging" in s]
    if not target_sheets:
        target_sheets = [wb.active.title]

    all_records = []
    report_date = None
    for sname in target_sheets:
        recs, rdate = parse_sheet(wb[sname])
        all_records.extend(recs)
        if rdate and not report_date:
            report_date = rdate

    records = all_records

    by_pic = {}
    for r in records:
        pic = str(r["pic"]).strip() if r["pic"] else "Chua phan cong"
        if pic not in by_pic:
            by_pic[pic] = []
        by_pic[pic].append(r)

    aging_order = {"Over 180 ngay": 0, "91-180 ngay": 1, "61-90 ngay": 2,
                   "31-60 ngay": 3, "1-30 ngay": 4, "Current": 5, "N/A": 6}
    for pic in by_pic:
        by_pic[pic].sort(key=lambda x: aging_order.get(x["aging_bucket"], 99))

    return {
        "report_date": report_date or "N/A",
        "total_records": len(records),
        "by_pic": by_pic
    }


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 parse_ar.py <path_to_xlsx>", file=sys.stderr)
        sys.exit(1)

    filepath = sys.argv[1]
    result = parse_ar_file(filepath)
    print(json.dumps(result, ensure_ascii=False, indent=2))
