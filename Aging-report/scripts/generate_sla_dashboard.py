#!/usr/bin/env python3
"""
Generate HTML SLA dashboard from AR data + tracker file.
Usage: python3 generate_sla_dashboard.py <ar_xlsx> [tracker_xlsx] <output.html>
"""

import sys, json, subprocess
from datetime import datetime, timedelta
from pathlib import Path

SCRIPT_DIR = str(Path(__file__).parent) + "/"

def add_working_days(start_date, days):
    if isinstance(start_date, str):
        try:
            start_date = datetime.strptime(start_date, "%d/%m/%Y")
        except:
            return None
    current = start_date
    added = 0
    while added < days:
        current += timedelta(days=1)
        if current.weekday() < 5:
            added += 1
    return current

def working_days_between(start, end):
    if not start or not end:
        return None
    if isinstance(start, str):
        try:
            start = datetime.strptime(start, "%d/%m/%Y")
        except:
            return None
    if isinstance(end, str):
        try:
            end = datetime.strptime(end, "%d/%m/%Y")
        except:
            return None
    days = 0
    current = start
    while current < end:
        current += timedelta(days=1)
        if current.weekday() < 5:
            days += 1
    return days

PAYMENT_SLA_DAYS = 30
PIC_RESPONSE_SLA_DAYS = 3

def payment_sla(over_day):
    if over_day <= 0:
        return "on-track", "✅ Chưa đến hạn"
    elif over_day <= PAYMENT_SLA_DAYS:
        return "at-risk", f"⚠️ Trễ {over_day} ngày"
    else:
        return "breached", f"🔴 Vi phạm ({over_day} ngày)"

def pic_sla(email_sent_str, response_str, today):
    if not email_sent_str:
        return "no-email", "— Chưa gửi mail"
    try:
        sent = datetime.strptime(email_sent_str, "%d/%m/%Y")
    except:
        return "no-email", "— Chưa gửi mail"
    if response_str:
        try:
            responded = datetime.strptime(response_str, "%d/%m/%Y")
            wd = working_days_between(sent, responded)
            if wd is not None and wd <= PIC_RESPONSE_SLA_DAYS:
                return "on-track", f"✅ Phản hồi ({wd} ngày)"
            else:
                return "at-risk", f"⚠️ Phản hồi muộn ({wd} ngày)"
        except:
            pass
    wd_elapsed = working_days_between(sent, today)
    if wd_elapsed is None:
        return "no-email", "— Chưa gửi mail"
    elif wd_elapsed <= PIC_RESPONSE_SLA_DAYS:
        deadline = add_working_days(sent, PIC_RESPONSE_SLA_DAYS)
        days_left = working_days_between(today, deadline) if today < deadline else 0
        return "at-risk", f"⏳ Còn {days_left} ngày làm việc"
    else:
        return "breached", f"🔴 Vi phạm ({wd_elapsed} ngày KLV)"

def load_tracking(tracker_xlsx):
    tracking = {}
    if not tracker_xlsx or not Path(tracker_xlsx).exists():
        return tracking
    try:
        from openpyxl import load_workbook
        wb = load_workbook(tracker_xlsx, data_only=True)
        if "Tracking" not in wb.sheetnames:
            return tracking
        ws = wb["Tracking"]
        headers = {}
        for row in ws.iter_rows(min_row=3, max_row=3, values_only=True):
            for i, h in enumerate(row):
                if h:
                    headers[str(h).strip()] = i
        for row in ws.iter_rows(min_row=5, values_only=True):
            inv_no = row[headers.get("Invoice #", 3)] if "Invoice #" in headers else None
            if not inv_no:
                continue
            email_sent = row[headers.get("Ngày gửi mail", 9)] if "Ngày gửi mail" in headers else None
            pic_resp = row[headers.get("Ngày PIC phản hồi", 10)] if "Ngày PIC phản hồi" in headers else None
            notes = row[headers.get("Ghi chú", 12)] if "Ghi chú" in headers else None
            key = str(inv_no).strip().lstrip("#")
            tracking[key] = {
                "email_sent": email_sent.strftime("%d/%m/%Y") if hasattr(email_sent, "strftime") else (str(email_sent) if email_sent else ""),
                "pic_response": pic_resp.strftime("%d/%m/%Y") if hasattr(pic_resp, "strftime") else (str(pic_resp) if pic_resp else ""),
                "notes": notes or "",
            }
    except Exception as e:
        print(f"Warning loading tracker: {e}", file=sys.stderr)
    return tracking

def generate_dashboard(ar_xlsx, tracker_xlsx, output_html):
    r = subprocess.run(
        [sys.executable, f"{SCRIPT_DIR}parse_ar.py", ar_xlsx],
        capture_output=True, text=True
    )
    if r.returncode != 0:
        raise RuntimeError(f"parse_ar.py failed: {r.stderr}")
    data = json.loads(r.stdout)
    report_date = data["report_date"]
    by_pic = data["by_pic"]
    today = datetime.now()
    tracking = load_tracking(tracker_xlsx)

    # Build rows
    rows = []
    for pic, records in by_pic.items():
        for rec in records:
            inv_key = str(rec["invoice_no"]) if rec["invoice_no"] else ""
            prev = tracking.get(inv_key, {})
            email_sent = prev.get("email_sent", "")
            pic_resp = prev.get("pic_response", "")
            notes = prev.get("notes", "")

            pay_cls, pay_label = payment_sla(rec["over_day"])
            pic_cls, pic_label = pic_sla(email_sent, pic_resp, today)

            rows.append({
                "pic": pic,
                "code": rec["code"],
                "name": rec["name"] or "",
                "invoice": f"#{rec['invoice_no']}" if rec["invoice_no"] else "N/A",
                "invoice_date": rec["invoice_date"] or "",
                "description": (rec["description"] or "").strip(),
                "amount": rec["base_amount"],
                "over_day": rec["over_day"],
                "aging_bucket": rec["aging_bucket"],
                "email_sent": email_sent,
                "pic_response": pic_resp,
                "notes": notes,
                "pay_cls": pay_cls,
                "pay_label": pay_label,
                "pic_cls": pic_cls,
                "pic_label": pic_label,
            })

    # KPI counts
    total = len(rows)
    pay_ok = sum(1 for r in rows if r["pay_cls"] == "on-track")
    pay_risk = sum(1 for r in rows if r["pay_cls"] == "at-risk")
    pay_breach = sum(1 for r in rows if r["pay_cls"] == "breached")
    pic_ok = sum(1 for r in rows if r["pic_cls"] == "on-track")
    pic_risk = sum(1 for r in rows if r["pic_cls"] == "at-risk")
    pic_breach = sum(1 for r in rows if r["pic_cls"] == "breached")
    pic_no_email = sum(1 for r in rows if r["pic_cls"] == "no-email")

    total_amount = sum(r["amount"] for r in rows)
    breach_amount = sum(r["amount"] for r in rows if r["pay_cls"] == "breached")

    rows_json = json.dumps(rows, ensure_ascii=False)
    pics = sorted(set(r["pic"] for r in rows))
    pic_options = "".join(f'<option value="{p}">{p}</option>' for p in pics)

    html = f"""<!DOCTYPE html>
<html lang="vi">
<head>
<meta charset="UTF-8">
<title>SLA Dashboard — AR Aged Debtor</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: Arial, sans-serif; background: #F0F4F8; color: #1a1a1a; font-size: 14px; }}
  .header {{ background: #1F4E79; color: white; padding: 18px 32px; display: flex; justify-content: space-between; align-items: center; }}
  .header h1 {{ font-size: 20px; font-weight: 700; }}
  .header .meta {{ font-size: 12px; opacity: 0.85; text-align: right; line-height: 1.6; }}
  .content {{ padding: 24px 32px; }}

  /* KPI cards */
  .kpi-section {{ margin-bottom: 28px; }}
  .kpi-section h2 {{ font-size: 13px; font-weight: 700; color: #555; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 12px; }}
  .kpi-row {{ display: flex; gap: 14px; flex-wrap: wrap; }}
  .kpi-card {{ background: white; border-radius: 10px; padding: 16px 20px; min-width: 140px; flex: 1; box-shadow: 0 1px 4px rgba(0,0,0,0.08); border-top: 4px solid #ccc; }}
  .kpi-card.green {{ border-top-color: #4CAF50; }}
  .kpi-card.yellow {{ border-top-color: #FFC107; }}
  .kpi-card.red {{ border-top-color: #F44336; }}
  .kpi-card.blue {{ border-top-color: #2196F3; }}
  .kpi-card.gray {{ border-top-color: #9E9E9E; }}
  .kpi-card .num {{ font-size: 32px; font-weight: 700; line-height: 1.1; }}
  .kpi-card .label {{ font-size: 12px; color: #666; margin-top: 4px; }}
  .kpi-card.green .num {{ color: #2E7D32; }}
  .kpi-card.yellow .num {{ color: #E65100; }}
  .kpi-card.red .num {{ color: #C62828; }}
  .kpi-card.blue .num {{ color: #0D47A1; }}
  .kpi-card.gray .num {{ color: #616161; }}

  /* Filters */
  .filter-bar {{ background: white; border-radius: 10px; padding: 14px 20px; margin-bottom: 18px; display: flex; gap: 14px; flex-wrap: wrap; align-items: center; box-shadow: 0 1px 4px rgba(0,0,0,0.07); }}
  .filter-bar label {{ font-size: 12px; color: #555; font-weight: 600; }}
  .filter-bar select, .filter-bar input {{ border: 1px solid #ddd; border-radius: 6px; padding: 6px 10px; font-size: 13px; color: #333; background: #fafafa; }}
  .filter-bar select:focus, .filter-bar input:focus {{ outline: none; border-color: #2196F3; }}
  .reset-btn {{ background: #E3F2FD; color: #1565C0; border: none; border-radius: 6px; padding: 6px 14px; font-size: 13px; cursor: pointer; font-weight: 600; }}
  .reset-btn:hover {{ background: #BBDEFB; }}

  /* Table */
  .table-wrap {{ background: white; border-radius: 10px; box-shadow: 0 1px 4px rgba(0,0,0,0.07); overflow: hidden; }}
  .table-info {{ padding: 12px 20px; font-size: 13px; color: #555; border-bottom: 1px solid #eee; }}
  table {{ width: 100%; border-collapse: collapse; }}
  thead th {{ background: #1F4E79; color: white; padding: 11px 12px; text-align: left; font-size: 12px; font-weight: 600; white-space: nowrap; cursor: pointer; user-select: none; }}
  thead th:hover {{ background: #2E75B6; }}
  tbody tr {{ border-bottom: 1px solid #f0f0f0; transition: background 0.15s; }}
  tbody tr:hover {{ background: #F8FAFF; }}
  tbody td {{ padding: 10px 12px; font-size: 13px; vertical-align: middle; }}
  .amount {{ text-align: right; font-variant-numeric: tabular-nums; }}

  /* SLA badges */
  .badge {{ display: inline-block; padding: 3px 10px; border-radius: 20px; font-size: 11px; font-weight: 700; white-space: nowrap; }}
  .badge.on-track {{ background: #E8F5E9; color: #2E7D32; }}
  .badge.at-risk {{ background: #FFF9C4; color: #E65100; }}
  .badge.breached {{ background: #FFEBEE; color: #C62828; }}
  .badge.no-email {{ background: #F5F5F5; color: #757575; }}

  .footer {{ text-align: center; padding: 20px; color: #999; font-size: 11px; }}
</style>
</head>
<body>

<div class="header">
  <div>
    <h1>📊 SLA Dashboard — AR Aged Debtor</h1>
    <div style="font-size:13px;opacity:0.85;margin-top:4px;">Theo dõi Payment SLA (NET{PAYMENT_SLA_DAYS}) & PIC Response SLA ({PIC_RESPONSE_SLA_DAYS} ngày làm việc)</div>
  </div>
  <div class="meta">
    Báo cáo đến: <strong>{report_date}</strong><br>
    Xem lúc: <strong>{today.strftime("%d/%m/%Y %H:%M")}</strong>
  </div>
</div>

<div class="content">

  <!-- Payment SLA KPIs -->
  <div class="kpi-section">
    <h2>💳 Payment SLA (NET{PAYMENT_SLA_DAYS})</h2>
    <div class="kpi-row">
      <div class="kpi-card blue"><div class="num">{total}</div><div class="label">Tổng invoice</div></div>
      <div class="kpi-card green"><div class="num">{pay_ok}</div><div class="label">✅ Chưa đến hạn</div></div>
      <div class="kpi-card yellow"><div class="num">{pay_risk}</div><div class="label">⚠️ Đang trễ (trong SLA)</div></div>
      <div class="kpi-card red"><div class="num">{pay_breach}</div><div class="label">🔴 Vi phạm SLA</div></div>
      <div class="kpi-card red"><div class="num">{int(breach_amount):,}</div><div class="label">VND vi phạm</div></div>
    </div>
  </div>

  <!-- PIC Response SLA KPIs -->
  <div class="kpi-section">
    <h2>📬 PIC Response SLA ({PIC_RESPONSE_SLA_DAYS} ngày làm việc)</h2>
    <div class="kpi-row">
      <div class="kpi-card green"><div class="num">{pic_ok}</div><div class="label">✅ Đã phản hồi đúng hạn</div></div>
      <div class="kpi-card yellow"><div class="num">{pic_risk}</div><div class="label">⏳ Đang chờ phản hồi</div></div>
      <div class="kpi-card red"><div class="num">{pic_breach}</div><div class="label">🔴 Vi phạm / Không phản hồi</div></div>
      <div class="kpi-card gray"><div class="num">{pic_no_email}</div><div class="label">— Chưa gửi mail</div></div>
    </div>
  </div>

  <!-- Filters -->
  <div class="filter-bar">
    <label>PIC:</label>
    <select id="f-pic" onchange="applyFilters()">
      <option value="">Tất cả</option>
      {pic_options}
    </select>
    <label>Payment SLA:</label>
    <select id="f-pay" onchange="applyFilters()">
      <option value="">Tất cả</option>
      <option value="on-track">✅ Chưa đến hạn</option>
      <option value="at-risk">⚠️ Đang trễ</option>
      <option value="breached">🔴 Vi phạm</option>
    </select>
    <label>PIC Response:</label>
    <select id="f-pic-sla" onchange="applyFilters()">
      <option value="">Tất cả</option>
      <option value="on-track">✅ Đã phản hồi</option>
      <option value="at-risk">⏳ Đang chờ</option>
      <option value="breached">🔴 Vi phạm</option>
      <option value="no-email">— Chưa gửi</option>
    </select>
    <label>Tìm kiếm:</label>
    <input type="text" id="f-search" placeholder="Invoice, KH, mô tả..." oninput="applyFilters()" style="width:180px">
    <button class="reset-btn" onclick="resetFilters()">↺ Reset</button>
  </div>

  <!-- Table -->
  <div class="table-wrap">
    <div class="table-info" id="table-info">Đang tải...</div>
    <table id="main-table">
      <thead>
        <tr>
          <th onclick="sortBy(0)">PIC ↕</th>
          <th onclick="sortBy(1)">Khách hàng ↕</th>
          <th onclick="sortBy(2)">Invoice ↕</th>
          <th onclick="sortBy(3)">Ngày HĐ ↕</th>
          <th onclick="sortBy(4)">Mô tả</th>
          <th onclick="sortBy(5)" style="text-align:right">Số tiền (VND) ↕</th>
          <th onclick="sortBy(6)">Quá hạn ↕</th>
          <th onclick="sortBy(7)">Payment SLA ↕</th>
          <th onclick="sortBy(8)">Ngày gửi mail ↕</th>
          <th onclick="sortBy(9)">PIC Response SLA ↕</th>
          <th>Ghi chú</th>
        </tr>
      </thead>
      <tbody id="table-body"></tbody>
    </table>
  </div>

</div>

<div class="footer">Cập nhật bằng cách upload file AR mới mỗi tháng — SLA Dashboard tự tính lại</div>

<script>
const ALL_ROWS = {rows_json};
let sortCol = -1, sortAsc = true;

function fmt(n) {{
  return n.toLocaleString('vi-VN');
}}

function render(rows) {{
  const tbody = document.getElementById('table-body');
  tbody.innerHTML = rows.map(r => `
    <tr>
      <td><strong>${{r.pic}}</strong></td>
      <td>${{r.code}}${{r.name ? ' — ' + r.name : ''}}</td>
      <td>${{r.invoice}}</td>
      <td>${{r.invoice_date}}</td>
      <td style="max-width:220px;word-break:break-word">${{r.description}}</td>
      <td class="amount">${{fmt(r.amount)}}</td>
      <td style="text-align:center">${{r.over_day > 0 ? r.over_day + ' ngày' : '—'}}</td>
      <td><span class="badge ${{r.pay_cls}}">${{r.pay_label}}</span></td>
      <td style="text-align:center">${{r.email_sent || '—'}}</td>
      <td><span class="badge ${{r.pic_cls}}">${{r.pic_label}}</span></td>
      <td style="color:#666;font-size:12px">${{r.notes || ''}}</td>
    </tr>
  `).join('');
  document.getElementById('table-info').textContent =
    `Hiển thị ${{rows.length}} / ${{ALL_ROWS.length}} invoice`;
}}

function applyFilters() {{
  const fPic = document.getElementById('f-pic').value;
  const fPay = document.getElementById('f-pay').value;
  const fPicSla = document.getElementById('f-pic-sla').value;
  const fSearch = document.getElementById('f-search').value.toLowerCase();
  let rows = ALL_ROWS.filter(r =>
    (!fPic || r.pic === fPic) &&
    (!fPay || r.pay_cls === fPay) &&
    (!fPicSla || r.pic_cls === fPicSla) &&
    (!fSearch || [r.invoice, r.code+'', r.name, r.description].join(' ').toLowerCase().includes(fSearch))
  );
  if (sortCol >= 0) rows = sortRows(rows);
  render(rows);
}}

function resetFilters() {{
  ['f-pic','f-pay','f-pic-sla'].forEach(id => document.getElementById(id).value = '');
  document.getElementById('f-search').value = '';
  applyFilters();
}}

const SORT_KEYS = ['pic','code','invoice','invoice_date','description','amount','over_day','pay_cls','email_sent','pic_cls'];
function sortRows(rows) {{
  const key = SORT_KEYS[sortCol];
  return [...rows].sort((a, b) => {{
    let av = a[key], bv = b[key];
    if (typeof av === 'number') return sortAsc ? av - bv : bv - av;
    return sortAsc ? String(av).localeCompare(String(bv)) : String(bv).localeCompare(String(av));
  }});
}}
function sortBy(col) {{
  if (sortCol === col) sortAsc = !sortAsc; else {{ sortCol = col; sortAsc = true; }}
  applyFilters();
}}

applyFilters();
</script>
</body>
</html>"""

    Path(output_html).write_text(html, encoding="utf-8")
    return output_html

if __name__ == "__main__":
    args = sys.argv[1:]
    if len(args) == 2:
        ar_xlsx, output_html = args
        tracker_xlsx = None
    elif len(args) == 3:
        ar_xlsx, tracker_xlsx, output_html = args
    else:
        print("Usage: python3 generate_sla_dashboard.py <ar.xlsx> [tracker.xlsx] <output.html>", file=sys.stderr)
        sys.exit(1)
    generate_dashboard(ar_xlsx, tracker_xlsx, output_html)
    print(f"Dashboard saved: {output_html}")
